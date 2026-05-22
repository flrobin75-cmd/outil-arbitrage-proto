"""
Génère les valeurs cibles Excel pour les cas 2-6 du test de parité TNS.

Procédure :
1. Charge le classeur original
2. Injecte les inputs (Profil + module TNS)
3. Sauvegarde une copie
4. Recalcule via LibreOffice headless
5. Extrait les valeurs cibles du module TNS recalculé
"""

import subprocess
import os
import shutil
from openpyxl import load_workbook


def injecter_inputs(input_path, output_path,
                    forme_juridique="SARL (gérance majoritaire) / EURL",
                    effectif="11-49 salariés",
                    situation="Marié / pacsé",
                    parts=2.0,
                    situation_part="Aucune (cas général)",
                    autres_rev=0,
                    div_foyer=0,
                    capital_cca=100_000,
                    rem_nette=70_000,
                    frais_reels=0,
                    div_bruts=50_000):
    """Copie le classeur et injecte les inputs."""
    shutil.copy(input_path, output_path)
    wb = load_workbook(output_path)

    # Profil (5. Profil)
    ws_profil = wb['5. Profil']
    ws_profil['C12'] = forme_juridique
    ws_profil['C13'] = ("TNS" if "majoritaire" in forme_juridique or "EI" in forme_juridique
                        else "TNS (libéral)" if "BNC" in forme_juridique or "SELARL" in forme_juridique
                        else "Assimilé salarié")
    ws_profil['C14'] = effectif
    # C15 part détenue - non utilisée pour ce module
    ws_profil['C16'] = capital_cca  # On met aussi dans le Profil pour cohérence
    ws_profil['C19'] = situation
    ws_profil['C20'] = parts
    ws_profil['C21'] = autres_rev
    ws_profil['C22'] = div_foyer
    ws_profil['C23'] = situation_part

    # Module TNS (8. TNS)
    ws_tns = wb['8. TNS']
    ws_tns['C5'] = rem_nette
    ws_tns['C6'] = frais_reels
    ws_tns['C42'] = capital_cca  # Module local
    ws_tns['C44'] = div_bruts

    wb.save(output_path)


def recalculer_libreoffice(xlsx_path):
    """Force le recalcul via LibreOffice headless."""
    # Convertir en xlsx puis re-sauver pour forcer recalcul
    result = subprocess.run(
        ['libreoffice', '--headless', '--calc', '--convert-to', 'xlsx',
         '--outdir', os.path.dirname(xlsx_path), xlsx_path],
        capture_output=True, timeout=60
    )
    return result.returncode == 0


def extraire_resultats_tns(xlsx_path):
    """Lit les résultats du module TNS recalculé."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['8. TNS']
    return {
        "Cotisations TNS totales (C10)": ws['C10'].value,
        "CSG déductible (C11)": ws['C11'].value,
        "CSG/CRDS non déd. (C12)": ws['C12'].value,
        "Revenu net pro (C15)": ws['C15'].value,
        "Revenu imposable (C16)": ws['C16'].value,
        "Revenu imposable foyer (C17)": ws['C17'].value,
        "Revenu par part (C18)": ws['C18'].value,
        "IR par part (C19)": ws['C19'].value,
        "IR avec QF (C20)": ws['C20'].value,
        "IR sans QF (C21)": ws['C21'].value,
        "Plafonnement QF (C22)": ws['C22'].value,
        "IR foyer après plafond QF (C23)": ws['C23'].value,
        "CEHR (C25)": ws['C25'].value,
        "CDHR (C27)": ws['C27'].value,
        "Total impôts foyer (C28)": ws['C28'].value,
        "Taux moyen IR (C29)": ws['C29'].value,
        "Coût total société (C38)": ws['C38'].value,
        "Seuil 10 % div (C43)": ws['C43'].value,
        "Fraction cotis TNS (C45)": ws['C45'].value,
        "Cotis TNS sur div (C46)": ws['C46'].value,
        "Fraction PFU (C47)": ws['C47'].value,
        "PFU sur fraction (C48)": ws['C48'].value,
        "IR sur fraction TNS (C49)": ws['C49'].value,
        "Net dividendes (C50)": ws['C50'].value,
    }


CAS = {
    "Cas 2 - Célibataire 1 part TMI 30%": dict(
        forme_juridique="SARL (gérance majoritaire) / EURL",
        situation="Célibataire / divorcé / veuf",
        parts=1.0,
        autres_rev=0,
        div_foyer=0,
        capital_cca=50_000,
        rem_nette=60_000,
        frais_reels=0,
        div_bruts=10_000,
    ),
    "Cas 3 - Marié 4 parts (2 enfants) revenu élevé": dict(
        forme_juridique="SARL (gérance majoritaire) / EURL",
        situation="Marié / pacsé",
        parts=4.0,
        autres_rev=20_000,
        div_foyer=0,
        capital_cca=200_000,
        rem_nette=150_000,
        frais_reels=2_000,
        div_bruts=30_000,
    ),
    "Cas 4 - Foyer riche déclenchant CEHR": dict(
        forme_juridique="SARL (gérance majoritaire) / EURL",
        situation="Marié / pacsé",
        parts=2.0,
        autres_rev=0,
        div_foyer=50_000,
        capital_cca=500_000,
        rem_nette=400_000,
        frais_reels=0,
        div_bruts=0,
    ),
    "Cas 5 - CDHR plancher 20%": dict(
        forme_juridique="SARL (gérance majoritaire) / EURL",
        situation="Célibataire / divorcé / veuf",
        parts=1.0,
        autres_rev=0,
        div_foyer=200_000,
        capital_cca=100_000,
        rem_nette=180_000,
        frais_reels=0,
        div_bruts=0,
    ),
    "Cas 6 - Dividendes sous le seuil 10%": dict(
        forme_juridique="SARL (gérance majoritaire) / EURL",
        situation="Marié / pacsé",
        parts=3.0,
        autres_rev=0,
        div_foyer=0,
        capital_cca=300_000,
        rem_nette=80_000,
        frais_reels=0,
        div_bruts=20_000,  # < 10% de 300 000 = 30 000, donc tout au PFU
    ),
}


if __name__ == "__main__":
    base = '/home/claude/outil_v19.xlsx'
    work_dir = '/home/claude/tns_dev/cibles'
    os.makedirs(work_dir, exist_ok=True)

    resultats_tous = {}

    for nom_cas, inputs in CAS.items():
        print(f"\n--- Génération cibles pour : {nom_cas} ---")
        work_file = os.path.join(work_dir, f"{nom_cas.replace(' ', '_').replace('%','pct')}.xlsx")
        injecter_inputs(base, work_file, **inputs)
        ok = recalculer_libreoffice(work_file)
        if not ok:
            print("  ✗ Recalcul LibreOffice échoué")
            continue
        # LibreOffice écrase le fichier dans outdir, donc le path est le même
        resultats = extraire_resultats_tns(work_file)
        resultats_tous[nom_cas] = resultats
        for label, val in resultats.items():
            print(f"    {label:50s} = {val}")

    # Sauvegarder en JSON pour réutilisation
    import json
    with open('/home/claude/tns_dev/cibles_excel.json', 'w') as f:
        json.dump(resultats_tous, f, indent=2, default=str)
    print(f"\n✓ Cibles sauvegardées dans cibles_excel.json")

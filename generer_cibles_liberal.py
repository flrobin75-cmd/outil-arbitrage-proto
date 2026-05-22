"""
Génère les cibles Excel pour les cas 2-6 du module Libéral.
Réutilise le pattern validé sur TNS.
"""

import subprocess
import os
import shutil
import json
from openpyxl import load_workbook


def injecter_inputs_liberal(input_path, output_path,
                            forme_juridique="Profession libérale (BNC)",
                            effectif="Sans salarié",
                            situation="Marié / pacsé",
                            parts=2.0,
                            situation_part="Aucune (cas général)",
                            autres_rev=0,
                            div_foyer=0,
                            recettes=150_000,
                            frais_pro=30_000,
                            benefice_sel=200_000,
                            rem_sel=80_000):
    """Copie + injection des inputs Profil + Module Libéral."""
    shutil.copy(input_path, output_path)
    wb = load_workbook(output_path)

    # Profil
    ws_profil = wb['5. Profil']
    ws_profil['C12'] = forme_juridique
    ws_profil['C13'] = "TNS (libéral)"  # régime social libéral
    ws_profil['C14'] = effectif
    ws_profil['C19'] = situation
    ws_profil['C20'] = parts
    ws_profil['C21'] = autres_rev
    ws_profil['C22'] = div_foyer
    ws_profil['C23'] = situation_part

    # Module Libéral
    ws_lib = wb['9. Libéral']
    ws_lib['C5'] = recettes
    ws_lib['C6'] = frais_pro
    ws_lib['C28'] = benefice_sel
    ws_lib['C29'] = rem_sel

    wb.save(output_path)


def extraire_resultats_liberal(xlsx_path):
    """Lit toutes les valeurs cibles du module Libéral."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['9. Libéral']
    return {
        "Bénéfice BNC (C7)": ws['C7'].value,
        "Cotisations URSSAF (C8)": ws['C8'].value,
        "CSG/CRDS non déd. (C9)": ws['C9'].value,
        "Bénéfice net après cotis. (C10)": ws['C10'].value,
        "Revenu imposable libéral (C11)": ws['C11'].value,
        "Revenu imposable foyer (C12)": ws['C12'].value,
        "Revenu par part (C13)": ws['C13'].value,
        "IR avec QF (C15)": ws['C15'].value,
        "IR sans QF (C16)": ws['C16'].value,
        "Plafonnement QF (C17)": ws['C17'].value,
        "IR foyer après plafond QF (C18)": ws['C18'].value,
        "CEHR (C20)": ws['C20'].value,
        "CDHR (C22)": ws['C22'].value,
        "Total impôts foyer (C23)": ws['C23'].value,
        "Impôts imputables libéral (C24)": ws['C24'].value,
        "Net libéral après impôts (C25)": ws['C25'].value,
        "Bénéfice imposable IS (C30)": ws['C30'].value,
        "IS dû (C31)": ws['C31'].value,
        "Résultat net distribuable (C32)": ws['C32'].value,
        "Dividendes envisagés (C33)": ws['C33'].value,
    }


CAS_LIBERAL = {
    "Cas 2 - BNC faible revenu célibataire 1 part": dict(
        situation="Célibataire / divorcé / veuf", parts=1.0,
        autres_rev=0, div_foyer=0,
        recettes=80_000, frais_pro=15_000,
        benefice_sel=200_000, rem_sel=80_000,
    ),
    "Cas 3 - BNC marié 3 parts avec autres revenus": dict(
        situation="Marié / pacsé", parts=3.0,
        autres_rev=30_000, div_foyer=0,
        recettes=200_000, frais_pro=40_000,
        benefice_sel=200_000, rem_sel=80_000,
    ),
    "Cas 4 - BNC haut revenu marié déclenchant CEHR": dict(
        situation="Marié / pacsé", parts=2.0,
        autres_rev=0, div_foyer=100_000,
        recettes=600_000, frais_pro=80_000,
        benefice_sel=200_000, rem_sel=80_000,
    ),
    "Cas 5 - BNC célibataire CDHR potentielle": dict(
        situation="Célibataire / divorcé / veuf", parts=1.0,
        autres_rev=0, div_foyer=250_000,
        recettes=500_000, frais_pro=50_000,
        benefice_sel=200_000, rem_sel=80_000,
    ),
    "Cas 6 - SEL double couche bénéfice élevé": dict(
        situation="Marié / pacsé", parts=2.0,
        autres_rev=0, div_foyer=0,
        recettes=150_000, frais_pro=30_000,
        benefice_sel=500_000, rem_sel=120_000,
    ),
}


if __name__ == "__main__":
    base = '/home/claude/outil_v19.xlsx'
    work_dir = '/home/claude/tns_dev/cibles_liberal'
    os.makedirs(work_dir, exist_ok=True)

    recalc_dir = '/tmp/recalc_lo_liberal'
    os.makedirs(recalc_dir, exist_ok=True)

    resultats_tous = {}

    for nom_cas, inputs in CAS_LIBERAL.items():
        print(f"\n--- {nom_cas} ---")
        safe = nom_cas.replace(' ', '_').replace('%', 'pct')
        src = os.path.join(work_dir, f"{safe}.xlsx")
        injecter_inputs_liberal(base, src, **inputs)

        out = os.path.join(recalc_dir, os.path.basename(src))
        if os.path.exists(out):
            os.remove(out)

        result = subprocess.run(
            ['libreoffice', '--headless', '--calc', '--convert-to', 'xlsx',
             '--outdir', recalc_dir, src],
            capture_output=True, timeout=120
        )

        if not os.path.exists(out):
            print(f"  ✗ Recalcul échoué")
            continue

        res = extraire_resultats_liberal(out)
        resultats_tous[nom_cas] = res
        for k, v in res.items():
            if v is not None:
                if isinstance(v, float):
                    print(f"  {k:50s} = {v:>15,.4f}")
                else:
                    print(f"  {k:50s} = {v}")

    with open('/home/claude/tns_dev/cibles_liberal.json', 'w') as f:
        json.dump(resultats_tous, f, indent=2, default=str)
    print(f"\n✓ Cibles sauvegardées")

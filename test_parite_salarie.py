"""
Génère cibles Excel + teste parité Salarié.
"""
import subprocess, os, shutil, json, sys
sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import load_workbook
from core.profil import Profil
from regime.salarie import calcul_module_salarie


def injecter_inputs_salarie(input_path, output_path,
                            forme_juridique="SAS / SASU",
                            effectif="11-49 salariés",
                            situation="Marié / pacsé",
                            parts=2.0,
                            situation_part="Aucune (cas général)",
                            autres_rev=0,
                            div_foyer=0,
                            salaire_brut=60_000):
    shutil.copy(input_path, output_path)
    wb = load_workbook(output_path)

    ws_profil = wb['5. Profil']
    ws_profil['C12'] = forme_juridique
    ws_profil['C13'] = "Salarié"
    ws_profil['C14'] = effectif
    ws_profil['C19'] = situation
    ws_profil['C20'] = parts
    ws_profil['C21'] = autres_rev
    ws_profil['C22'] = div_foyer
    ws_profil['C23'] = situation_part

    ws_sal = wb['10. Salarié']
    ws_sal['C5'] = salaire_brut

    wb.save(output_path)


def extraire_resultats_salarie(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['10. Salarié']
    return {
        "Cotis salariales (C8)": ws['C8'].value,
        "CSG/CRDS totale (C9)": ws['C9'].value,
        "CSG déductible (C10)": ws['C10'].value,
        "Net avant impôt (C11)": ws['C11'].value,
        "Revenu salarial imposable (C14)": ws['C14'].value,
        "Abattement 10% (C15)": ws['C15'].value,
        "Revenu imposable net (C16)": ws['C16'].value,
        "Revenu imposable foyer (C17)": ws['C17'].value,
        "IR foyer après plafond QF (C23)": ws['C23'].value,
        "CEHR (C25)": ws['C25'].value,
        "CDHR (C27)": ws['C27'].value,
        "Total impôts foyer (C28)": ws['C28'].value,
        "Impôts imputables rém (C29)": ws['C29'].value,
        "Net après impôts (C30)": ws['C30'].value,
    }


CAS = {
    "Cas 1 - défaut v19 (brut 60k marié 2 parts)": dict(
        situation="Marié / pacsé", parts=2.0, autres_rev=0, div_foyer=0,
        salaire_brut=60_000,
    ),
    "Cas 2 - Célibataire 1 part bas salaire": dict(
        situation="Célibataire / divorcé / veuf", parts=1.0,
        autres_rev=0, div_foyer=0, salaire_brut=35_000,
    ),
    "Cas 3 - Marié 4 parts salaire moyen avec autres revenus": dict(
        situation="Marié / pacsé", parts=4.0,
        autres_rev=25_000, div_foyer=0, salaire_brut=80_000,
    ),
    "Cas 4 - Cadre supérieur abattement plafonné": dict(
        situation="Marié / pacsé", parts=2.0,
        autres_rev=0, div_foyer=0, salaire_brut=200_000,
    ),
    "Cas 5 - Très haut salaire célibataire CEHR": dict(
        situation="Célibataire / divorcé / veuf", parts=1.0,
        autres_rev=0, div_foyer=150_000, salaire_brut=350_000,
    ),
    "Cas 6 - Très très haut salaire couple CDHR potentielle": dict(
        situation="Marié / pacsé", parts=2.0,
        autres_rev=0, div_foyer=400_000, salaire_brut=150_000,
    ),
}


# --- Génération des cibles Excel ---
base = '/home/claude/outil_v19.xlsx'
work_dir = '/home/claude/tns_dev/cibles_salarie'
recalc_dir = '/tmp/recalc_lo_salarie'
os.makedirs(work_dir, exist_ok=True)
os.makedirs(recalc_dir, exist_ok=True)

cibles_tous = {}
for nom, inputs in CAS.items():
    safe = nom.replace(' ', '_').replace('%', 'pct').replace('(', '').replace(')', '')
    src = os.path.join(work_dir, f"{safe}.xlsx")
    injecter_inputs_salarie(base, src, **inputs)

    out = os.path.join(recalc_dir, os.path.basename(src))
    if os.path.exists(out):
        os.remove(out)

    subprocess.run(
        ['libreoffice', '--headless', '--calc', '--convert-to', 'xlsx',
         '--outdir', recalc_dir, src],
        capture_output=True, timeout=120,
    )

    if os.path.exists(out):
        cibles_tous[nom] = extraire_resultats_salarie(out)

with open('/home/claude/tns_dev/cibles_salarie.json', 'w') as f:
    json.dump(cibles_tous, f, indent=2, default=str)

# --- Tests de parité ---
def fmt(v):
    if v is None: return "       None"
    if isinstance(v, (int, float)): return f"{v:>15,.4f}"
    return f"{str(v):>15s}"


def comparer(label, py, xl, tol=0.01):
    if xl is None: return None, f"  ⚠ {label:55s}"
    if py is None: return False, f"  ✗ {label:55s} Python=None  Excel={fmt(xl)}"
    ecart = py - xl
    ok = abs(ecart) <= tol
    marker = "✓" if ok else "✗"
    return ok, f"  {marker} {label:55s} Python={fmt(py)}  Excel={fmt(xl)}  écart={ecart:+10,.4f}"


total_ok = 0
total = 0
resume = []

for nom, inputs in CAS.items():
    print("=" * 110)
    print(f"  {nom}")
    print("=" * 110)
    profil = Profil(
        forme_juridique="SAS / SASU",
        effectif="11-49 salariés",
        situation=inputs["situation"],
        parts=inputs["parts"],
        autres_revenus=inputs["autres_rev"],
        dividendes_foyer_hors_enveloppe=inputs["div_foyer"],
    )
    res = calcul_module_salarie(profil, salaire_brut=inputs["salaire_brut"])

    correspondances = {
        "Cotis salariales (C8)": res.cotis_salariales,
        "CSG/CRDS totale (C9)": res.csg_crds_totale,
        "CSG déductible (C10)": res.csg_deductible,
        "Net avant impôt (C11)": res.net_avant_impot,
        "Revenu salarial imposable (C14)": res.revenu_salarial_imposable,
        "Abattement 10% (C15)": res.abattement_10pct,
        "Revenu imposable net (C16)": res.revenu_imposable_net,
        "Revenu imposable foyer (C17)": res.revenu_imposable_foyer,
        "IR foyer après plafond QF (C23)": res.ir_foyer,
        "CEHR (C25)": res.cehr,
        "CDHR (C27)": res.cdhr,
        "Total impôts foyer (C28)": res.total_impots_foyer,
        "Impôts imputables rém (C29)": res.impots_imputables_rem,
        "Net après impôts (C30)": res.net_apres_impots,
    }

    cibles = cibles_tous.get(nom, {})
    ok_cas, total_cas = 0, 0
    for label, py in correspondances.items():
        xl = cibles.get(label)
        if isinstance(xl, str):
            try: xl = float(xl)
            except: pass
        ok, ligne = comparer(label, py, xl)
        print(ligne)
        if ok is not None:
            total_cas += 1
            if ok: ok_cas += 1
    print(f"\n  Résultat : {ok_cas}/{total_cas} OK\n")
    resume.append((nom, ok_cas, total_cas))
    total_ok += ok_cas
    total += total_cas

print("=" * 110)
print("  SYNTHÈSE SALARIÉ")
print("=" * 110)
for nom, ok, t in resume:
    status = "✓" if ok == t else "✗"
    print(f"  {status} {nom:60s} : {ok}/{t}")
print(f"\n  TOTAL : {total_ok}/{total} cellules en parité")

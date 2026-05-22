"""
Tests Scénarios A vs B - Mini doctrine QA formalisée.

Cas 1 : Stabilité du calcul (2 scénarios identiques → écart = 0)
Cas 2 : Validation arbitrage dividendes (avant/après distribution)
Cas 3 : Validation inter-régimes (Assimilé vs TNS)
Cas 4 : Robustesse variation enveloppe (sensibilité salaire)

Tolérance : 0,01 € (cohérence avec autres modules).
Génération des cibles via injection Excel + recalcul LibreOffice.
"""

import subprocess, os, shutil, json, sys
sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import load_workbook
from strategy.scenarios import ScenarioInputs, calcul_comparaison


def injecter_scenarios(input_path, output_path,
                       scenario_a: dict, scenario_b: dict,
                       situation="Marié / pacsé"):
    """Copie + injection des inputs Scénarios A vs B."""
    shutil.copy(input_path, output_path)
    wb = load_workbook(output_path)

    # Profil pour cohérence (situation)
    ws_profil = wb['5. Profil']
    ws_profil['C19'] = situation

    # Module Scénarios
    ws_sc = wb['Scénarios A vs B']
    # Scénario A
    ws_sc['C6'] = scenario_a.get("libelle", "Cas A")
    ws_sc['C7'] = scenario_a.get("situation", situation)
    ws_sc['C8'] = scenario_a.get("parts", 2.0)
    ws_sc['C9'] = scenario_a.get("regime", "Assimilé salarié")
    ws_sc['C10'] = scenario_a.get("salaire", 100000)
    ws_sc['C11'] = scenario_a.get("dividendes", 0)
    ws_sc['C12'] = scenario_a.get("epargne", 0)
    ws_sc['C13'] = scenario_a.get("peripheriques", 0)
    # Scénario B
    ws_sc['D6'] = scenario_b.get("libelle", "Cas B")
    ws_sc['D7'] = scenario_b.get("situation", situation)
    ws_sc['D8'] = scenario_b.get("parts", 2.0)
    ws_sc['D9'] = scenario_b.get("regime", "Assimilé salarié")
    ws_sc['D10'] = scenario_b.get("salaire", 100000)
    ws_sc['D11'] = scenario_b.get("dividendes", 0)
    ws_sc['D12'] = scenario_b.get("epargne", 0)
    ws_sc['D13'] = scenario_b.get("peripheriques", 0)

    wb.save(output_path)


def extraire_resultats_scenarios(xlsx_path):
    """Extrait les résultats des 2 scénarios + projection."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['Scénarios A vs B']

    return {
        # Scénario A
        ("A", "net_salaire_apres_cotis"): ws['C17'].value,
        ("A", "revenu_imposable"): ws['C18'].value,
        ("A", "ir_barème"): ws['C19'].value,
        ("A", "net_salaire_apres_ir"): ws['C20'].value,
        ("A", "net_dividendes"): ws['C21'].value,
        ("A", "net_epargne"): ws['C22'].value,
        ("A", "net_peripheriques"): ws['C23'].value,
        ("A", "total_net"): ws['C25'].value,
        ("A", "proj_y1"): ws['C31'].value,
        ("A", "proj_y5"): ws['C35'].value,
        # Scénario B
        ("B", "net_salaire_apres_cotis"): ws['D17'].value,
        ("B", "revenu_imposable"): ws['D18'].value,
        ("B", "ir_barème"): ws['D19'].value,
        ("B", "net_salaire_apres_ir"): ws['D20'].value,
        ("B", "net_dividendes"): ws['D21'].value,
        ("B", "net_epargne"): ws['D22'].value,
        ("B", "net_peripheriques"): ws['D23'].value,
        ("B", "total_net"): ws['D25'].value,
        ("B", "proj_y1"): ws['D31'].value,
        ("B", "proj_y5"): ws['D35'].value,
        # Écart
        ("ecart", "total"): ws['E25'].value,
        ("ecart", "pourcent"): ws['E26'].value,
    }


def fmt(v):
    if v is None: return "None"
    if isinstance(v, (int, float)): return f"{v:,.4f}".replace(",", " ")
    return str(v)


def comparer(label, py, xl, tol=0.01):
    if xl is None: return None
    if isinstance(xl, str):
        try: xl = float(xl)
        except: return None
    if py is None: return False
    return abs(py - xl) <= tol


# ============================================================
# DÉFINITION DES 4 CAS DOCTRINAUX
# ============================================================
CAS_DOCTRINE = {
    "Cas 1 - Stabilité (2 scénarios identiques)": {
        "objectif": "Stabilité du calcul - écart attendu = 0",
        "scenario_a": dict(libelle="Cas de base", parts=2.0,
                           regime="Assimilé salarié", salaire=100000,
                           dividendes=0, epargne=0, peripheriques=0),
        "scenario_b": dict(libelle="Identique", parts=2.0,
                           regime="Assimilé salarié", salaire=100000,
                           dividendes=0, epargne=0, peripheriques=0),
    },
    "Cas 2 - Arbitrage dividendes (avant/après distribution)": {
        "objectif": "Validation arbitrage dividendes",
        "scenario_a": dict(libelle="Sans dividendes", parts=2.0,
                           regime="Assimilé salarié", salaire=100000,
                           dividendes=0, epargne=0, peripheriques=0),
        "scenario_b": dict(libelle="Avec dividendes 50k", parts=2.0,
                           regime="Assimilé salarié", salaire=100000,
                           dividendes=50000, epargne=15000, peripheriques=4000),
    },
    "Cas 3 - Inter-régimes (Assimilé vs TNS)": {
        "objectif": "Validation différence de régimes",
        "scenario_a": dict(libelle="Assimilé salarié", parts=2.0,
                           regime="Assimilé salarié", salaire=80000,
                           dividendes=20000, epargne=0, peripheriques=0),
        "scenario_b": dict(libelle="TNS gérant majoritaire", parts=2.0,
                           regime="TNS", salaire=80000,
                           dividendes=20000, epargne=0, peripheriques=0),
    },
    "Cas 4 - Sensibilité enveloppe (variation salaire)": {
        "objectif": "Robustesse variation enveloppe",
        "scenario_a": dict(libelle="Enveloppe 100k", parts=2.0,
                           regime="Assimilé salarié", salaire=100000,
                           dividendes=10000, epargne=5000, peripheriques=2000),
        "scenario_b": dict(libelle="Enveloppe 150k", parts=2.0,
                           regime="Assimilé salarié", salaire=150000,
                           dividendes=10000, epargne=5000, peripheriques=2000),
    },
}


# ============================================================
# GÉNÉRATION CIBLES + COMPARAISON
# ============================================================
base = '/home/claude/outil_v19.xlsx'
work_dir = '/home/claude/tns_dev/cibles_scenarios'
recalc_dir = '/tmp/recalc_lo_scenarios'
os.makedirs(work_dir, exist_ok=True)
os.makedirs(recalc_dir, exist_ok=True)


cibles_tous = {}
resume = []
total_ok = 0
total = 0

for nom_cas, defi in CAS_DOCTRINE.items():
    print("=" * 110)
    print(f"  {nom_cas}")
    print(f"  Objectif : {defi['objectif']}")
    print("=" * 110)

    # Génération cible Excel
    safe = nom_cas.replace(' ', '_').replace('/', '_').replace('(', '').replace(')', '')
    src = os.path.join(work_dir, f"{safe}.xlsx")
    injecter_scenarios(base, src, defi["scenario_a"], defi["scenario_b"])

    out = os.path.join(recalc_dir, os.path.basename(src))
    if os.path.exists(out):
        os.remove(out)

    subprocess.run(
        ['libreoffice', '--headless', '--calc', '--convert-to', 'xlsx',
         '--outdir', recalc_dir, src],
        capture_output=True, timeout=120,
    )

    if not os.path.exists(out):
        print(f"  ✗ Recalcul Excel échoué")
        continue

    cibles_excel = extraire_resultats_scenarios(out)
    cibles_tous[nom_cas] = {f"{k[0]}.{k[1]}": v for k, v in cibles_excel.items()}

    # Calcul Python - mapping des clés vers ScenarioInputs
    def _to_inputs(d):
        return ScenarioInputs(
            libelle=d.get("libelle", "Cas"),
            situation=d.get("situation", "Marié / pacsé"),
            parts=d.get("parts", 2.0),
            regime_social=d.get("regime", "Assimilé salarié"),
            salaire_brut=d.get("salaire", 0),
            dividendes_bruts=d.get("dividendes", 0),
            epargne_salariale_per=d.get("epargne", 0),
            peripheriques=d.get("peripheriques", 0),
        )
    sc_a = _to_inputs(defi["scenario_a"])
    sc_b = _to_inputs(defi["scenario_b"])
    res = calcul_comparaison(sc_a, sc_b)

    valeurs_python = {
        ("A", "net_salaire_apres_cotis"): res.scenario_a.net_salaire_apres_cotis,
        ("A", "revenu_imposable"): res.scenario_a.revenu_imposable,
        ("A", "ir_barème"): res.scenario_a.ir_barème,
        ("A", "net_salaire_apres_ir"): res.scenario_a.net_salaire_apres_ir,
        ("A", "net_dividendes"): res.scenario_a.net_dividendes,
        ("A", "net_epargne"): res.scenario_a.net_epargne,
        ("A", "net_peripheriques"): res.scenario_a.net_peripheriques,
        ("A", "total_net"): res.scenario_a.total_net,
        ("A", "proj_y1"): res.scenario_a.projection_5_ans[0],
        ("A", "proj_y5"): res.scenario_a.projection_5_ans[4],
        ("B", "net_salaire_apres_cotis"): res.scenario_b.net_salaire_apres_cotis,
        ("B", "revenu_imposable"): res.scenario_b.revenu_imposable,
        ("B", "ir_barème"): res.scenario_b.ir_barème,
        ("B", "net_salaire_apres_ir"): res.scenario_b.net_salaire_apres_ir,
        ("B", "net_dividendes"): res.scenario_b.net_dividendes,
        ("B", "net_epargne"): res.scenario_b.net_epargne,
        ("B", "net_peripheriques"): res.scenario_b.net_peripheriques,
        ("B", "total_net"): res.scenario_b.total_net,
        ("B", "proj_y1"): res.scenario_b.projection_5_ans[0],
        ("B", "proj_y5"): res.scenario_b.projection_5_ans[4],
        ("ecart", "total"): res.ecart_total,
        ("ecart", "pourcent"): res.ecart_pourcent,
    }

    ok_cas, total_cas = 0, 0
    for key, py_val in valeurs_python.items():
        xl_val = cibles_excel.get(key)
        ok = comparer(key, py_val, xl_val)
        if ok is None:
            continue
        marker = "✓" if ok else "✗"
        label = f"{key[0]}.{key[1]}"
        if isinstance(xl_val, str):
            try: xl_val = float(xl_val)
            except: pass
        ecart_val = py_val - xl_val if isinstance(xl_val, (int, float)) else 0
        print(f"  {marker} {label:35s} Python={fmt(py_val):>15s}  "
              f"Excel={fmt(xl_val):>15s}  écart={ecart_val:+.4f}")
        total_cas += 1
        if ok: ok_cas += 1

    print(f"\n  Résultat : {ok_cas}/{total_cas} OK\n")
    resume.append((nom_cas, ok_cas, total_cas))
    total_ok += ok_cas
    total += total_cas


# Sauvegarde des cibles pour versionnage (recommandation gouvernance)
with open('/home/claude/tns_dev/cibles_scenarios.json', 'w') as f:
    json.dump(cibles_tous, f, indent=2, default=str)

print("=" * 110)
print("  SYNTHÈSE TESTS SCÉNARIOS A vs B")
print("=" * 110)
for nom, ok, t in resume:
    status = "✓" if ok == t else "✗"
    print(f"  {status} {nom:60s} : {ok}/{t}")
print(f"\n  TOTAL : {total_ok}/{total} cellules en parité v19 stricte")
print(f"  Cibles versionnées dans : /home/claude/tns_dev/cibles_scenarios.json")
